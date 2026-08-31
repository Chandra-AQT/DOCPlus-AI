"""
excel_service.py — Excel report generator for DocPlus discovery results.
"""
import os
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DOC_TYPE_LABELS = {
    "PSS": "Product Specification Sheet",
    "IOM": "Installation, Operations & Maintenance",
    "OWN": "Owner's Manual / User Guide",
    "SVM": "Service Manual / Technical Manual",
    "SVB": "Service Bulletins",
    "PCT": "Product Catalog",
    "PBR": "Product Brochure",
    "SUB": "Submittal",
    "WDG": "Wiring Diagram",
    "PLD": "Parts List & Exploded Diagram",
    "WTY": "Warranty Statement",
    "CCL": "Commissioning Checklist",
    "RCL": "Recall Notices",
    "SDS": "Safety Data Sheet",
    "CRT": "Compliance & Certification",
    "RUG": "Retrofit & Upgrade Guides",
    "OTHER": "Other / Unclassified",
}
FORMAT_LABELS = {"pdf": "PDF", "word": "Word", "excel": "Excel", "ppt": "PowerPoint"}


def create_excel(items, folder: str) -> str:
    excel_path = os.path.join(folder, "pdf_links.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    header_fill = PatternFill("solid", fgColor="0284C7")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")
    headers = ["#", "Filename", "Format", "Document Type", "Type Description", "Domain", "URL"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, item in enumerate(items, start=1):
        if isinstance(item, str):
            url = item
            filename = url.split("/")[-1].split("?")[0].strip() or "document"
            fmt = "pdf"
            doc_type = "OTHER"
        else:
            url = item.get("url", "")
            filename = item.get("filename", url.split("/")[-1].split("?")[0] or "document")
            fmt = item.get("format", "pdf")
            doc_type = item.get("doc_type", "OTHER")
        parsed = urlparse(url)
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=filename)
        ws.cell(row=i+1, column=3, value=FORMAT_LABELS.get(fmt, fmt.upper()))
        ws.cell(row=i+1, column=4, value=doc_type)
        ws.cell(row=i+1, column=5, value=DOC_TYPE_LABELS.get(doc_type, ""))
        ws.cell(row=i+1, column=6, value=parsed.netloc)
        ws.cell(row=i+1, column=7, value=url)
    col_widths = [6, 45, 14, 10, 40, 30, 90]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    wb.save(excel_path)
    return excel_path
